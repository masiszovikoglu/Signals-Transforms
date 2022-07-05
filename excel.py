# Reading an excel file using Python

from pkgutil import get_data
import pandas as pd

import os

import openpyxl
import numpy as np
 
# Give the location of the file
path = "water.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

# Dictionary to hold the date value pairs
my_dict = {"Date":[],"Value":[]};
 
# number of coulmns of excel file
max_col = sheet_obj.max_column
# number of rows of excel file
m_row = sheet_obj.max_row

for i in range(2, m_row + 1):
    cell_obj = sheet_obj.cell(row = i, column = 1)
    x=cell_obj.value.split()
    line=x[12].split(";")
    date=line[4]+" "+ line[5]
    height=int(line[7])
    my_dict["Date"].append(date)
    my_dict["Value"].append(height)
# for i in range(len(my_dict["Date"])):    
#     print ("Date"+ my_dict["Date"][i] +" = " + my_dict["Value"][i])   
for i in range(len(my_dict["Value"])):
    if my_dict["Value"][i] > 1000000:
        j = i
        while True:
            j += 1
            if my_dict["Value"][j] <= 1000000:
                break
        arr = np.arange(my_dict["Value"][i-1], my_dict["Value"][j+1], (my_dict["Value"][j+1] - my_dict["Value"][i-1]) / (j - i + 2))
        for k in range(i, j):
            my_dict["Value"][k] = int(arr[k + 1 - i])

def get_data():
    return my_dict

path="C:\\Users\\masis\\Downloads\\20220704_065.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

# Dictionary to hold the date value pairs
my_dict2 = {"Date":[],"Value":[]};
 
# number of coulmns of excel file
max_col = sheet_obj.max_column
# number of rows of excel file
m_row = sheet_obj.max_row

for i in range(2, m_row + 1):
    cell_obj = sheet_obj.cell(row = i, column = 1)
    x=cell_obj.value.split()
    line=x[12].split(";")
    date=line[4]+" "+ line[5]
    height=int(line[7])
    my_dict2["Date"].append(date)
    my_dict2["Value"].append(height)
# for i in range(len(my_dict["Date"])):    
#     print ("Date"+ my_dict["Date"][i] +" = " + my_dict["Value"][i])   
for i in range(len(my_dict2["Value"])):
    if my_dict2["Value"][i] > 1000000:
        j = i
        while True:
            j += 1
            if my_dict2["Value"][j] <= 1000000:
                break
        arr = np.arange(my_dict2["Value"][i-1], my_dict2["Value"][j+1], (my_dict2["Value"][j+1] - my_dict2["Value"][i-1]) / (j - i + 2))
        for k in range(i, j):
            my_dict2["Value"][k] = int(arr[k + 1 - i])
def get_data_now():
    return my_dict2            